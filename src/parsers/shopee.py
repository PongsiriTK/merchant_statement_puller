"""ShopeeFood daily sales report parser.

Expected file format: .xlsx with columns:
    A: No.
    B: Transaction Type
    C: Order ID
    ...
    I: Order Amount
    N: Total
    O: Commission
    P: VAT on Commission
    T: Net Income
    U: Order Status

Email subject pattern:
    "{merchant_name} รายงานการโอนเงินสำหรับ ShopeeFood {YYYY-MM-DD}"

Attachment filename pattern:
    "01_รายงานยอดขายประจำวัน_{merchant_name}_{DDMMMYY}.xlsx"
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from src.core.models import DailyReport, OrderRecord
from .base import BaseParser

logger = logging.getLogger(__name__)

# Column indices (0-based) in the ShopeeFood daily report
COL_ORDER_ID = 2       # C: Order ID
COL_STORE_NAME = 5     # F: Store Name
COL_ORDER_TIME = 6     # G: Order Create Time
COL_ORDER_AMOUNT = 8   # I: Order Amount
COL_TOTAL = 13         # N: Total
COL_COMMISSION = 14    # O: Commission
COL_VAT = 15           # P: VAT on Commission
COL_NET_INCOME = 19    # T: Net Income
COL_STATUS = 20        # U: Order Status


def _to_decimal(value) -> Decimal:
    """Safely convert a cell value to Decimal."""
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _extract_date_from_filename(filename: str) -> date:
    """Extract report date from filename like '..._20Mar26.xlsx'.

    The date format is DDMMMYY where MMM is English month abbreviation.
    """
    match = re.search(r"_(\d{2})([A-Za-z]{3})(\d{2})\.", filename)
    if match:
        day, month_str, year_short = match.groups()
        date_str = f"{day} {month_str} {year_short}"
        try:
            return datetime.strptime(date_str, "%d %b %y").date()
        except ValueError:
            pass

    # Fallback: try to extract from email subject (YYYY-MM-DD)
    logger.warning("Could not extract date from filename: %s", filename)
    return date.today()


def _extract_date_from_subject(subject: str) -> date | None:
    """Try to extract YYYY-MM-DD date from email subject."""
    match = re.search(r"(\d{4}-\d{2}-\d{2})", subject)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


class ShopeeParser(BaseParser):
    """Parser for ShopeeFood daily sales report .xlsx files."""

    platform_name = "ShopeeFood"

    def build_search_subject(self, merchant_name: str) -> str:
        """ShopeeFood emails use:
        '{merchant_name} รายงานการโอนเงินสำหรับ ShopeeFood'
        """
        return f"{merchant_name} รายงานการโอนเงินสำหรับ ShopeeFood"

    def attachment_filename_filter(self, merchant_name: str) -> str:
        """Filter for the daily sales report attachment."""
        return "รายงานยอดขายประจำวัน"

    def parse_file(
        self, filepath: Path, email_subject: str = "", email_date: str = "",
        password: str = "",
    ) -> DailyReport:
        """Parse a ShopeeFood .xlsx daily sales report.

        Reads row 1 as header, then data rows from row 2 onward.
        Sums the 'Total' column (N) for all settled orders.
        """
        logger.info("Parsing ShopeeFood report: %s", filepath.name)

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Determine report date
        report_date = _extract_date_from_filename(filepath.name)
        if report_date == date.today() and email_subject:
            extracted = _extract_date_from_subject(email_subject)
            if extracted:
                report_date = extracted

        orders: list[OrderRecord] = []
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

        for row in rows:
            if not row or row[0] is None:
                continue

            order_id = str(row[COL_ORDER_ID] or "")
            store_name = str(row[COL_STORE_NAME] or "")
            order_time = row[COL_ORDER_TIME]

            # Parse order date from the order create time
            order_date = report_date
            if order_time:
                try:
                    order_date = datetime.strptime(
                        str(order_time)[:10], "%Y-%m-%d"
                    ).date()
                except (ValueError, TypeError):
                    pass

            orders.append(
                OrderRecord(
                    order_id=order_id,
                    order_date=order_date,
                    order_amount=_to_decimal(row[COL_ORDER_AMOUNT]),
                    total=_to_decimal(row[COL_TOTAL]),
                    net_income=_to_decimal(row[COL_NET_INCOME]),
                    commission=_to_decimal(row[COL_COMMISSION]),
                    vat_on_commission=_to_decimal(row[COL_VAT]),
                    status=str(row[COL_STATUS] or "unknown"),
                    store_name=store_name,
                )
            )

        wb.close()

        report = DailyReport(
            report_date=report_date,
            filename=filepath.name,
            orders=orders,
            email_subject=email_subject,
            email_date=email_date,
        )

        logger.info(
            "Parsed %d orders, Total=%.2f, NetIncome=%.2f",
            report.order_count,
            report.total_sales,
            report.total_net_income,
        )

        return report
