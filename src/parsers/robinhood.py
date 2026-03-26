"""Parser for Robinhood (โรบินฮู้ด) daily merchant reports.

Robinhood sends password-protected .xlsx files as daily settlement reports.
The file is an OLE2-encrypted OOXML workbook that must be decrypted with
``msoffcrypto`` before ``openpyxl`` can read it.

Sheet layout (Sheet: "Statement"):
    Row  6 : Company header ("Purple Ventures CO.,LTD")
    Row 14 : Report date in cell C14 (format: DD-Mon-YY, e.g. "17-Feb-26")
    Row 17 : Column headers
    Row 18+: Order data rows
    Row N  : Summary row where col J == "ยอดรวมทั้งหมด" (grand total)

Key columns (1-indexed):
    B  (2)  : Shop ID
    C  (3)  : Shop Name
    D  (4)  : Order no.
    E  (5)  : Transaction Type
    I  (9)  : Order Time
    J  (10) : Order Completed Time
    K  (11) : จำนวนเงิน (Order Amount)
    L  (12) : ส่วนลดค่าอาหาร (Food Discount)
    M  (13) : ค่าสินค้าสุทธิ (Net Product Value) ← PRIMARY COLUMN
    Q  (17) : Processing Fee
    R  (18) : VAT on Processing Fee
    U  (21) : Net
    X  (24) : Withholding Tax
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

import msoffcrypto
import openpyxl

from src.core.models import DailyReport, OrderRecord
from src.parsers.base import BaseParser

logger = logging.getLogger(__name__)

# --- Column indices (1-based, matching openpyxl) ---
COL_SHOP_ID = 2        # B
COL_SHOP_NAME = 3      # C
COL_ORDER_NO = 4       # D
COL_TXN_TYPE = 5       # E
COL_ORDER_TIME = 9     # I
COL_COMPLETED_TIME = 10 # J
COL_ORDER_AMOUNT = 11  # K  จำนวนเงิน
COL_FOOD_DISCOUNT = 12 # L  ส่วนลดค่าอาหาร
COL_NET_PRODUCT = 13   # M  ค่าสินค้าสุทธิ
COL_PROC_FEE = 17      # Q  Processing Fee
COL_VAT_PROC = 18      # R  VAT on Processing fee
COL_NET = 21           # U  Net
COL_WHT = 24           # X  Withholding Tax

HEADER_ROW = 17
DATA_START_ROW = 18
DATE_CELL_ROW = 14
DATE_CELL_COL = 3      # C14

SUMMARY_MARKER = "ยอดรวมทั้งหมด"


def _to_decimal(value) -> Decimal:
    """Safely convert a cell value to Decimal."""
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _decrypt_workbook(filepath: Path, password: str) -> io.BytesIO:
    """Decrypt a password-protected Office file and return an in-memory stream.

    Robinhood wraps the .xlsx in OLE2 encryption. We use ``msoffcrypto``
    to strip the encryption layer, producing a standard OOXML stream that
    ``openpyxl`` can open.

    Raises:
        ValueError: If the file cannot be decrypted with the given password.
    """
    with open(filepath, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)

        if not office_file.is_encrypted():
            # Not encrypted — return raw bytes
            f.seek(0)
            buf = io.BytesIO(f.read())
            return buf

        office_file.load_key(password=password)
        decrypted = io.BytesIO()

        try:
            office_file.decrypt(decrypted)
        except Exception as exc:
            raise ValueError(
                f"Failed to decrypt {filepath.name} — wrong password? ({exc})"
            ) from exc

    decrypted.seek(0)
    return decrypted


def _parse_report_date(raw_date) -> date:
    """Parse the date from cell C14.

    Robinhood uses formats like:
        "17-Feb-26"  (DD-Mon-YY)
        datetime objects (if Excel stored it as a date)
    """
    if isinstance(raw_date, datetime):
        return raw_date.date()
    if isinstance(raw_date, date):
        return raw_date

    text = str(raw_date).strip().strip("'")

    # Try DD-Mon-YY  (e.g. "17-Feb-26")
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Cannot parse Robinhood report date: {raw_date!r}")


class RobinhoodParser(BaseParser):
    """Parser for Robinhood daily merchant settlement reports."""

    platform_name = "Robinhood"

    @property
    def requires_file_password(self) -> bool:
        return True

    def build_search_subject(self, merchant_name: str) -> str:
        """Robinhood uses a fixed English subject line for all merchants."""
        return "Robinhood Shop :: Daily Merchant Report"

    def attachment_filename_filter(self, merchant_name: str) -> str:
        """Robinhood filenames start with 'STM' (Statement)."""
        return "STM"

    def parse_file(
        self,
        filepath: Path,
        email_subject: str = "",
        email_date: str = "",
        password: str = "",
    ) -> DailyReport:
        """Parse a Robinhood daily settlement .xlsx file.

        Args:
            filepath: Path to the encrypted .xlsx file.
            email_subject: Original email subject (for fallback date parsing).
            email_date: Original email date header.
            password: Decryption password for the file.

        Returns:
            A DailyReport with all orders and computed totals.
        """
        # --- Decrypt ---
        decrypted_stream = _decrypt_workbook(filepath, password)

        wb = openpyxl.load_workbook(decrypted_stream, read_only=True, data_only=True)

        sheet_name = "Statement" if "Statement" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        # --- Extract report date from C14 ---
        raw_date = ws.cell(row=DATE_CELL_ROW, column=DATE_CELL_COL).value
        try:
            report_date = _parse_report_date(raw_date)
        except ValueError:
            logger.warning(
                "Could not parse date from C14 (%r), falling back to email date",
                raw_date,
            )
            # Fallback: try to extract from email_subject or email_date
            report_date = _fallback_date(email_subject, email_date)

        # --- Parse order rows ---
        orders: list[OrderRecord] = []

        for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=COL_WHT, values_only=False):
            # Stop at summary row
            completed_time_val = row[COL_COMPLETED_TIME - 1].value
            if completed_time_val and str(completed_time_val).strip() == SUMMARY_MARKER:
                break

            order_no = row[COL_ORDER_NO - 1].value
            if not order_no:
                continue  # skip empty rows

            order_amount = _to_decimal(row[COL_ORDER_AMOUNT - 1].value)
            net_product = _to_decimal(row[COL_NET_PRODUCT - 1].value)
            proc_fee = _to_decimal(row[COL_PROC_FEE - 1].value)
            vat_proc = _to_decimal(row[COL_VAT_PROC - 1].value)
            net_income = _to_decimal(row[COL_NET - 1].value)
            shop_name = str(row[COL_SHOP_NAME - 1].value or "")

            orders.append(
                OrderRecord(
                    order_id=str(order_no),
                    order_date=report_date,
                    order_amount=order_amount,
                    total=net_product,           # ค่าสินค้าสุทธิ
                    net_income=net_income,
                    commission=proc_fee,         # Processing Fee as commission
                    vat_on_commission=vat_proc,  # VAT on Processing Fee
                    status="settled",
                    store_name=shop_name,
                )
            )

        wb.close()

        return DailyReport(
            report_date=report_date,
            filename=filepath.name,
            orders=orders,
            email_subject=email_subject,
            email_date=email_date,
        )


def _fallback_date(email_subject: str, email_date: str) -> date:
    """Try to extract a date from the email metadata when C14 fails."""
    # Try YYYY-MM-DD anywhere in subject
    match = re.search(r"(\d{4}-\d{2}-\d{2})", email_subject)
    if match:
        return datetime.strptime(match.group(1), "%Y-%m-%d").date()

    # Try parsing email_date header (RFC 2822)
    if email_date:
        from email.utils import parsedate_to_datetime

        try:
            return parsedate_to_datetime(email_date).date()
        except Exception:
            pass

    raise ValueError("Cannot determine report date from email metadata")
